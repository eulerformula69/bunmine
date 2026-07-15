from io import BytesIO
from datetime import date

from openpyxl import load_workbook

from backend.services.vocabulary_report_model import build_report_rows, is_non_lexical_token, pick_sentence, plain_anki_text
from backend.services.vocabulary_report_service import safe_report_filename
from backend.services.vocabulary_report_service import VocabularyReportError
from backend.services.vocabulary_report_workbook import SUMMARY_HEADERS, create_workbook
from backend.app import create_app


def sample_cues():
    return [
        {"episodeId": 1, "episode": "Episode 1", "start": 1.0, "end": 2.0, "sentence": "猫を見た", "tokens": [
            {"surface": "見", "basic": "見る", "reading": "ミ", "pos": "動詞", "position": 3, "candidates": ["見", "見る"]}]},
        {"episodeId": 2, "episode": "Episode 2", "start": 3.0, "end": 4.0, "sentence": "猫を見る", "tokens": [
            {"surface": "見る", "basic": "見る", "reading": "ミル", "pos": "動詞", "position": 3, "candidates": ["見る"]}]},
    ]


def test_aggregation_status_filter_and_first_occurrence():
    summary, occurrences, totals = build_report_rows("Show", sample_cues(), {"見る": {"status": "new"}}, set(), {"new"})
    assert len(summary) == 1 and summary[0]["count"] == 2
    assert len(summary[0]["episodes"]) == 2 and summary[0]["sentence"] == "猫を見た"
    assert len(occurrences) == 2 and totals["new"] == 2
    assert build_report_rows("Show", sample_cues(), {"見る": {"status": "mature"}}, set(), {"new"})[1] == []


def test_unknown_known_basic_and_all_anki_statuses():
    token = sample_cues()[:1]
    for status in ["new", "learning", "young", "mature", "suspended"]:
        assert build_report_rows("S", token, {"見る": {"status": status}}, set(), {status})[1][0]["status"] == status
    assert build_report_rows("S", token, {}, set(), {"not_in_anki"})[1][0]["status"] == "not_in_anki"
    assert build_report_rows("S", token, {}, {"見る"}, {"known_basic"})[1][0]["status"] == "known_basic"


def test_anki_sentence_selection_and_html_cleanup():
    fields = {"Sentence": {"value": ""}, "Example": {"value": "<b>日本語&amp;</b>[sound:x.mp3]"}}
    assert pick_sentence(fields, ["Sentence", "Example"]) == "日本語&"
    assert pick_sentence({}, ["Sentence"]) == ""
    assert plain_anki_text("<div>A<br>B</div>") == "A B"


def test_valid_workbook_selected_sheets_and_headers():
    summary, occurrences, totals = build_report_rows("Show", sample_cues(), {"見る": {"status": "new"}}, set(), {"new"})
    stream = create_workbook(summary, occurrences, totals, {"summary": True, "occurrences": False, "statistics": True})
    workbook = load_workbook(BytesIO(stream.getvalue()))
    assert workbook.sheetnames == ["Vocabulary Summary", "Statistics"]
    assert [cell.value for cell in workbook["Vocabulary Summary"][1]] == SUMMARY_HEADERS
    assert workbook["Vocabulary Summary"].freeze_panes == "A2"


def test_safe_filename():
    filename = safe_report_filename('Frieren: <Season>/1')
    assert filename.endswith(f"_vocabulary_report_{date.today().isoformat()}.xlsx")
    assert not any(character in filename for character in '<>:"/\\|?*')


def test_download_endpoint_xlsx_mime_and_cleanup(monkeypatch, tmp_path):
    report = tmp_path / "report.xlsx"; report.write_bytes(b"xlsx")
    monkeypatch.setattr("backend.routes.vocabulary_report_routes.get_job", lambda _job_id: {
        "kind": "vocabulary-report", "status": "completed", "result": {"path": str(report), "filename": "report.xlsx"}})
    client = create_app().test_client()
    response = client.get("/library/vocabulary-report/job/download")
    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.close()
    assert not report.exists()


def test_analyzer_is_always_decoded_as_utf8(monkeypatch):
    import backend.services.vocabulary_report_service as service
    captured = {}
    monkeypatch.setattr(service, "_series_files", lambda _id: ({"title": "番組"}, [{"path": "字幕.srt"}]))
    monkeypatch.setattr(service, "read_known_anki_data", lambda: {"words": {}})
    monkeypatch.setattr(service, "read_anki_highlight_settings", lambda: {})
    monkeypatch.setattr(service, "known_basic_words_path", lambda: None)
    monkeypatch.setattr(service, "read_words_file", lambda _path: [])
    monkeypatch.setattr(service, "build_report_rows", lambda *args, **kwargs: ([], [], {}))
    monkeypatch.setattr(service, "create_workbook", lambda *args: BytesIO(b"xlsx"))
    def fake_run(*args, **kwargs):
        captured.update(kwargs)
        return type("Result", (), {"returncode": 0, "stdout": '[{"sentence":"日本語"}]', "stderr": ""})()
    monkeypatch.setattr(service.subprocess, "run", fake_run)
    service.generate_vocabulary_report(1, {"statuses": ["new"], "sheets": {"summary": True}})
    assert captured["encoding"] == "utf-8"
    assert captured["errors"] == "strict"


def test_empty_analyzer_output_has_clear_error(monkeypatch):
    import backend.services.vocabulary_report_service as service
    monkeypatch.setattr(service, "_series_files", lambda _id: ({"title": "Show"}, [{"path": "sub.srt"}]))
    monkeypatch.setattr(service.subprocess, "run", lambda *args, **kwargs: type("Result", (), {"returncode": 0, "stdout": "", "stderr": ""})())
    try:
        service.generate_vocabulary_report(1, {"statuses": ["new"], "sheets": {"summary": True}})
        assert False, "expected VocabularyReportError"
    except VocabularyReportError as error:
        assert str(error) == "Subtitle analyzer returned no data"


def test_particles_are_optional_and_disabled_by_default():
    cues = [{"episodeId": 1, "episode": "E1", "start": 0, "end": 1, "sentence": "のから", "tokens": [
        {"surface": "の", "basic": "の", "pos": "名詞", "position": 1, "candidates": ["の"]},
        {"surface": "から", "basic": "から", "pos": "助詞", "position": 2, "candidates": ["から"]},
    ]}]
    assert build_report_rows("S", cues, {}, set(), {"not_in_anki"})[1] == []
    included = build_report_rows("S", cues, {}, set(), {"not_in_anki"}, include_particles=True)[1]
    assert [row["word"] for row in included] == ["の", "から"]


def test_numbers_punctuation_and_special_symbols_are_always_ignored():
    for surface, pos, detail in [
        ("１", "名詞", "数"), ("２", "名詞", "数"), ("♪", "記号", "一般"),
        ("♪〜", "記号", "一般"), ("･", "記号", "一般"), (")）", "記号", "括弧閉"),
        ("!?", "unknown", ""),
    ]:
        assert is_non_lexical_token({"surface": surface, "pos": pos, "posDetail": detail})
        assert is_non_lexical_token({"surface": surface, "pos": pos, "posDetail": detail}, include_particles=True)


def test_auxiliary_and_requested_conversational_forms_are_optional():
    requested_forms = "だ た ない ます う れる さん どう まし あっ お たい っす なん じゃん もの あ たち ぬ ら く ちゃう うい られる おお ええ たく がる おっ ご しまう す では はっ ほら り ま おい".split()
    for surface in requested_forms:
        token = {"surface": surface, "pos": "名詞", "posDetail": "一般"}
        assert is_non_lexical_token(token)
        assert not is_non_lexical_token(token, include_auxiliary_forms=True)
    auxiliary = {"surface": "だ", "pos": "助動詞", "posDetail": "*"}
    assert is_non_lexical_token(auxiliary)
    assert not is_non_lexical_token(auxiliary, include_auxiliary_forms=True)


def test_auxiliary_toggle_changes_report_rows():
    cues = [{"episodeId": 1, "episode": "E1", "start": 0, "end": 1, "sentence": "だった", "tokens": [
        {"surface": "だ", "basic": "だ", "pos": "助動詞", "position": 1, "candidates": ["だ"]},
        {"surface": "た", "basic": "た", "pos": "助動詞", "position": 2, "candidates": ["た"]},
    ]}]
    assert build_report_rows("S", cues, {}, set(), {"not_in_anki"})[1] == []
    included = build_report_rows("S", cues, {}, set(), {"not_in_anki"}, include_auxiliary_forms=True)[1]
    assert [row["word"] for row in included] == ["だ", "た"]


def test_normalized_auxiliary_form_is_filtered_when_surface_differs():
    cues = [{"episodeId": 1, "episode": "E1", "start": 0, "end": 1, "sentence": "られ", "tokens": [
        {"surface": "られ", "basic": "られる", "pos": "動詞", "position": 1, "candidates": ["られ", "られる"]},
    ]}]
    assert build_report_rows("S", cues, {}, set(), {"not_in_anki"})[1] == []
