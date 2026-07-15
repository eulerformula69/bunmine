from io import BytesIO
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SUMMARY_HEADERS = ["Series", "Word", "Surface Forms", "Reading", "Part of Speech", "Status", "Series Occurrences", "Episode Count", "First Episode", "First Timestamp", "Series Sentence", "Anki Sentence", "Deck", "Anki Note ID", "Anki Card IDs", "Match Type", "Anki Match Count"]
OCCURRENCE_HEADERS = ["Series", "Episode", "Episode ID", "Timestamp Start", "Timestamp End", "Word", "Surface Form", "Reading", "Part of Speech", "Status", "Subtitle Sentence", "Anki Sentence", "Deck", "Anki Note ID", "Match Type"]


def _table_sheet(ws, headers, rows):
    ws.append(headers); ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows)+1)}"
    for cell in ws[1]: cell.font = Font(bold=True)
    for row in rows: ws.append(row)
    for index, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(index)].width = 42 if "Sentence" in header else min(24, max(12, len(header)+2))
        if "Sentence" in header:
            for cell in ws[get_column_letter(index)]: cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_workbook(summary, occurrences, totals, sheets):
    wb = Workbook(); wb.remove(wb.active)
    if sheets.get("summary"):
        ws = wb.create_sheet("Vocabulary Summary")
        _table_sheet(ws, SUMMARY_HEADERS, [[r["series"], r["word"], ", ".join(sorted(r["surfaces"])), r["reading"], r["pos"], r["status"], r["count"], len(r["episodes"]), r["episode"], r["start"], r["sentence"], r["anki_sentence"], r["deck"], r["note_id"], ", ".join(map(str, r["card_ids"])), r["match_type"], r["match_count"]] for r in summary])
    if sheets.get("occurrences"):
        ws = wb.create_sheet("Occurrences")
        _table_sheet(ws, OCCURRENCE_HEADERS, [[r["series"], r["episode"], r["episode_id"], r["start"], r["end"], r["word"], r["surface"], r["reading"], r["pos"], r["status"], r["sentence"], r["anki_sentence"], r["deck"], r["note_id"], r["match_type"]] for r in occurrences])
    if sheets.get("statistics"):
        ws = wb.create_sheet("Statistics")
        metrics = [("Total token occurrences", sum(totals.get(name, 0) for name in ["not_in_anki", "new", "learning", "young", "mature", "suspended", "known_basic"])), ("Unique normalized words", totals.get("__unique__", 0))]
        metrics += [(f"{name.replace('_', ' ').title()} unique words", totals.get(f"__unique_{name}", 0)) for name in ["not_in_anki", "new", "learning", "young", "mature", "suspended", "known_basic"]]
        metrics += [("Included unique words", len(summary)), ("Included token occurrences", len(occurrences))]
        _table_sheet(ws, ["Metric", "Value"], metrics)
        ws.append([]); ws.append(["Episode", "Unique Words", "Unknown", "New", "Learning", "Young", "Included Words", "Included Occurrences"])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
        episodes = {}
        for row in occurrences:
            item = episodes.setdefault(row["episode"], {"words": set(), "statuses": {}, "count": 0})
            item["words"].add(row["word"]); item["statuses"].setdefault(row["status"], set()).add(row["word"]); item["count"] += 1
        for name, item in episodes.items(): ws.append([name, len(item["words"]), len(item["statuses"].get("not_in_anki", set())), len(item["statuses"].get("new", set())), len(item["statuses"].get("learning", set())), len(item["statuses"].get("young", set())), len(item["words"]), item["count"]])
        ws.append([]); ws.append(["Word", "Status", "Occurrences", "Episode Count", "Example Sentence"])
        for cell in ws[ws.max_row]: cell.font = Font(bold=True)
        for row in summary[:100]: ws.append([row["word"], row["status"], row["count"], len(row["episodes"]), row["sentence"]])
    stream = BytesIO(); wb.save(stream); stream.seek(0); return stream
